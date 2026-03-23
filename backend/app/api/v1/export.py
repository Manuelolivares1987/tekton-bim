"""Export API — generate downloadable reports and data files."""

import io
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models.bim_wall import BimWall
from app.models.project import Project
from app.services.bim_wall_service import BimWallService

router = APIRouter()


@router.get("/project/{project_id}/cubicacion/excel")
def export_cubicacion_excel(project_id: int, db: Session = Depends(get_db)):
    """Export full project cubicación (material takeoff) as Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as e:
        raise HTTPException(500, "openpyxl not installed") from e

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    walls = db.query(BimWall).filter(BimWall.project_id == project_id).order_by(BimWall.label).all()
    if not walls:
        raise HTTPException(400, "No walls in project")

    svc = BimWallService(db)

    # Create workbook
    wb = openpyxl.Workbook()

    # ── Sheet 1: Resumen ──
    ws = wb.active
    ws.title = "Resumen"

    # Header styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws["A1"] = "TEKTON BIM — Cubicación de Proyecto"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Proyecto: {project.name}"
    ws["A3"] = f"Fecha: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A4"] = f"Muros: {len(walls)}"

    # Summary table
    row = 6
    headers = ["Muro", "Largo (m)", "Alto (m)", "Espesor (mm)", "Paneles", "Aberturas", "Área Muro (m²)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    total_area = 0
    total_panels = 0
    total_openings = 0

    for wall in walls:
        row += 1
        assembly = svc.get_wall_assembly(wall.id)
        area = (wall.length_mm / 1000) * (wall.height_mm / 1000)
        total_area += area
        total_panels += len(assembly["panels"])
        total_openings += len(assembly["openings"])

        values = [
            wall.label,
            round(wall.length_mm / 1000, 2),
            round(wall.height_mm / 1000, 2),
            wall.thickness_mm,
            len(assembly["panels"]),
            len(assembly["openings"]),
            round(area, 2),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_panels).font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_openings).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_area, 2)).font = Font(bold=True)

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    # ── Sheet 2: Paneles detallados ──
    ws2 = wb.create_sheet("Paneles")
    row = 1
    panel_headers = ["Muro", "Panel", "Ancho (mm)", "Alto (mm)", "Espesor (mm)", "Área (m²)", "Vano", "Tipo Vano"]
    for col, h in enumerate(panel_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for wall in walls:
        assembly = svc.get_wall_assembly(wall.id)
        for panel in assembly["panels"]:
            row += 1
            area = (panel["width_mm"] / 1000) * (panel["height_mm"] / 1000)
            values = [
                wall.label,
                panel["label"],
                panel["width_mm"],
                panel["height_mm"],
                panel["thickness_mm"],
                round(area, 3),
                "Sí" if panel.get("has_opening") else "No",
                panel.get("opening_type") or "",
            ]
            for col, v in enumerate(values, 1):
                ws2.cell(row=row, column=col, value=v).border = border

    for col in ws2.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    # ── Sheet 3: Entramado (framing) ──
    ws3 = wb.create_sheet("Entramado")
    row = 1
    framing_headers = ["Muro", "Tipo", "Largo (mm)", "Ancho (mm)", "Profundidad (mm)", "Madera", "Cantidad"]
    for col, h in enumerate(framing_headers, 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    framing_summary: dict[str, dict] = {}

    for wall in walls:
        assembly = svc.get_wall_assembly(wall.id)
        for member in assembly["framing"]:
            row += 1
            values = [
                wall.label,
                member["member_type"],
                member["length_mm"],
                member["width_mm"],
                member["depth_mm"],
                member["lumber_size"],
                member.get("quantity", 1),
            ]
            for col, v in enumerate(values, 1):
                ws3.cell(row=row, column=col, value=v).border = border

            # Accumulate summary
            mt = member["member_type"]
            if mt not in framing_summary:
                framing_summary[mt] = {"count": 0, "total_length_mm": 0}
            framing_summary[mt]["count"] += member.get("quantity", 1)
            framing_summary[mt]["total_length_mm"] += member["length_mm"] * member.get("quantity", 1)

    # Framing summary at bottom
    row += 2
    ws3.cell(row=row, column=1, value="RESUMEN ENTRAMADO").font = Font(bold=True, size=11)
    row += 1
    for mt, data in sorted(framing_summary.items()):
        ws3.cell(row=row, column=1, value=mt)
        ws3.cell(row=row, column=2, value=data["count"]).font = Font(bold=True)
        ws3.cell(row=row, column=3, value=f"{data['total_length_mm'] / 1000:.1f} m total")
        row += 1

    for col in ws3.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tekton-cubicacion-{project.name.replace(' ', '_')}-{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
