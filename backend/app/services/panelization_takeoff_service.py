"""Advanced panelization takeoff service.

Generates detailed material quantities at per-panel, per-wall, and project levels.
"""

import math
from io import BytesIO

from sqlalchemy.orm import Session

from app.core.sip_constants import SIP_FASTENER_RULES
from app.models.framing_member import FramingMember
from app.models.ifc_element import IfcElement
from app.models.material import Material
from app.models.panelization_result import PanelInstance, PanelizationResult
from app.models.sip_panel_config import SipPanelConfig


class PanelizationTakeoffService:
    """Generates detailed material quantities from panelization results."""

    def __init__(self, db: Session):
        self.db = db

    def per_panel_takeoff(self, project_id: int, waste_factor: float = 1.1) -> list[dict]:
        """Get material breakdown for each panel instance."""
        results = (
            self.db.query(PanelizationResult)
            .filter(PanelizationResult.project_id == project_id)
            .all()
        )

        panels_takeoff = []
        for result in results:
            element = self.db.query(IfcElement).filter(IfcElement.id == result.ifc_element_id).first()
            config = self._get_sip_config(result) if result.config_type == "sip" else None

            for instance in result.panel_instances:
                materials = self._calculate_panel_materials(instance, config, result.config_type, waste_factor)
                panels_takeoff.append({
                    "panel_label": instance.panel_label,
                    "panel_type": instance.panel_type,
                    "wall_name": element.name if element else None,
                    "storey": element.storey_name if element else None,
                    "width_mm": instance.width_mm,
                    "height_mm": instance.height_mm,
                    "area_m2": instance.area_m2,
                    "weight_kg": instance.weight_kg,
                    "materials": materials,
                })

        return panels_takeoff

    def per_wall_takeoff(self, project_id: int, waste_factor: float = 1.1) -> list[dict]:
        """Get material summary aggregated by wall."""
        results = (
            self.db.query(PanelizationResult)
            .filter(PanelizationResult.project_id == project_id)
            .all()
        )

        walls_takeoff = []
        for result in results:
            element = self.db.query(IfcElement).filter(IfcElement.id == result.ifc_element_id).first()
            config = self._get_sip_config(result) if result.config_type == "sip" else None

            # Aggregate materials across all panels in this wall
            wall_materials: dict[str, dict] = {}
            total_area = 0.0
            total_weight = 0.0

            for instance in result.panel_instances:
                panel_mats = self._calculate_panel_materials(instance, config, result.config_type, waste_factor)
                total_area += instance.area_m2
                total_weight += instance.weight_kg

                for mat in panel_mats:
                    key = mat["material_name"]
                    if key in wall_materials:
                        wall_materials[key]["quantity"] += mat["quantity"]
                        wall_materials[key]["cost"] += mat["cost"]
                    else:
                        wall_materials[key] = {**mat}

            # Add joint materials (splines between panels)
            joint_count = max(0, len(result.panel_instances) - 1)
            if config and result.config_type == "sip" and joint_count > 0:
                panel_height = result.panel_instances[0].height_mm if result.panel_instances else 2440
                spline_key = "Spline (juntas)"
                spline_material = self.db.query(Material).filter(Material.id == config.spline_material_id).first() if config.spline_material_id else None
                spline_length = joint_count * (panel_height / 1000)
                wall_materials[spline_key] = {
                    "material_name": spline_key,
                    "category": "connector",
                    "quantity": round(spline_length * waste_factor, 2),
                    "unit": "ml",
                    "cost": round(spline_length * waste_factor * (spline_material.cost_per_unit if spline_material else 2.5), 2),
                }

            walls_takeoff.append({
                "wall_name": element.name if element else f"Muro {result.ifc_element_id}",
                "storey": element.storey_name if element else None,
                "panel_count": len(result.panel_instances),
                "total_area_m2": round(total_area, 2),
                "total_weight_kg": round(total_weight, 2),
                "config_type": result.config_type,
                "materials": list(wall_materials.values()),
            })

        return walls_takeoff

    def project_summary(self, project_id: int, waste_factor: float = 1.1) -> dict:
        """Get project-wide material summary."""
        wall_takeoffs = self.per_wall_takeoff(project_id, waste_factor)

        # Aggregate by material
        total_materials: dict[str, dict] = {}
        total_panels = 0
        total_area = 0.0
        total_weight = 0.0
        by_storey: dict[str, dict] = {}

        for wall in wall_takeoffs:
            total_panels += wall["panel_count"]
            total_area += wall["total_area_m2"]
            total_weight += wall["total_weight_kg"]

            storey = wall["storey"] or "Sin piso"
            if storey not in by_storey:
                by_storey[storey] = {"panels": 0, "area_m2": 0, "weight_kg": 0}
            by_storey[storey]["panels"] += wall["panel_count"]
            by_storey[storey]["area_m2"] += wall["total_area_m2"]
            by_storey[storey]["weight_kg"] += wall["total_weight_kg"]

            for mat in wall["materials"]:
                key = mat["material_name"]
                if key in total_materials:
                    total_materials[key]["quantity"] += mat["quantity"]
                    total_materials[key]["cost"] += mat["cost"]
                else:
                    total_materials[key] = {**mat}

        # Round aggregated values
        for mat in total_materials.values():
            mat["quantity"] = round(mat["quantity"], 2)
            mat["cost"] = round(mat["cost"], 2)

        total_cost = sum(m["cost"] for m in total_materials.values())

        return {
            "project_id": project_id,
            "total_walls": len(wall_takeoffs),
            "total_panels": total_panels,
            "total_area_m2": round(total_area, 2),
            "total_weight_kg": round(total_weight, 2),
            "total_cost": round(total_cost, 2),
            "waste_factor": waste_factor,
            "by_storey": by_storey,
            "materials": list(total_materials.values()),
        }

    def export_excel(self, project_id: int, waste_factor: float = 1.1) -> BytesIO:
        """Export takeoff to Excel with multiple sheets."""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        summary = self.project_summary(project_id, waste_factor)
        per_wall = self.per_wall_takeoff(project_id, waste_factor)
        per_panel = self.per_panel_takeoff(project_id, waste_factor)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Sheet 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(["Resumen del Proyecto"])
        ws1.append([])
        ws1.append(["Muros totales", summary["total_walls"]])
        ws1.append(["Paneles totales", summary["total_panels"]])
        ws1.append(["Area total (m²)", summary["total_area_m2"]])
        ws1.append(["Peso total (kg)", summary["total_weight_kg"]])
        ws1.append(["Costo total", summary["total_cost"]])
        ws1.append(["Factor de desperdicio", summary["waste_factor"]])
        ws1.append([])
        ws1.append(["Material", "Categoría", "Cantidad", "Unidad", "Costo unitario", "Costo total"])
        style_header(ws1, 10, 6)
        for mat in summary["materials"]:
            ws1.append([
                mat["material_name"], mat["category"],
                mat["quantity"], mat["unit"],
                mat.get("cost_per_unit", 0), mat["cost"],
            ])

        # Sheet 2: Por Muro
        ws2 = wb.create_sheet("Por Muro")
        ws2.append(["Muro", "Piso", "Paneles", "Area (m²)", "Peso (kg)", "Tipo"])
        style_header(ws2, 1, 6)
        for wall in per_wall:
            ws2.append([
                wall["wall_name"], wall["storey"],
                wall["panel_count"], wall["total_area_m2"],
                wall["total_weight_kg"], wall["config_type"],
            ])

        # Sheet 3: Por Panel
        ws3 = wb.create_sheet("Por Panel")
        ws3.append(["Etiqueta", "Tipo", "Muro", "Piso", "Ancho (mm)", "Alto (mm)", "Area (m²)", "Peso (kg)"])
        style_header(ws3, 1, 8)
        for panel in per_panel:
            ws3.append([
                panel["panel_label"], panel["panel_type"],
                panel["wall_name"], panel["storey"],
                panel["width_mm"], panel["height_mm"],
                panel["area_m2"], panel["weight_kg"],
            ])

        # Sheet 4: Lista de Compras
        ws4 = wb.create_sheet("Lista de Compras")
        ws4.append(["Material", "Categoría", "Cantidad", "Unidad", "Costo unitario", "Costo total", "Proveedor", "Notas"])
        style_header(ws4, 1, 8)
        for mat in summary["materials"]:
            ws4.append([
                mat["material_name"], mat["category"],
                mat["quantity"], mat["unit"],
                mat.get("cost_per_unit", 0), mat["cost"],
                "", "",  # proveedor y notas en blanco para llenar
            ])

        # Auto-width columns
        for ws in [ws1, ws2, ws3, ws4]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _get_sip_config(self, result: PanelizationResult) -> SipPanelConfig | None:
        if result.sip_config_id:
            return self.db.query(SipPanelConfig).filter(SipPanelConfig.id == result.sip_config_id).first()
        return None

    def _calculate_panel_materials(
        self, instance: PanelInstance, config: SipPanelConfig | None,
        config_type: str, waste_factor: float
    ) -> list[dict]:
        """Calculate materials for a single panel instance."""
        materials = []

        if config_type == "sip" and config:
            materials = self._sip_panel_materials(instance, config, waste_factor)
        elif config_type == "wood_frame":
            materials = self._wood_frame_panel_materials(instance, waste_factor)

        return materials

    def _sip_panel_materials(
        self, instance: PanelInstance, config: SipPanelConfig, waste_factor: float
    ) -> list[dict]:
        """Calculate materials for a SIP panel."""
        area = instance.area_m2
        materials = []

        # Skin material (both faces)
        skin_mat = self.db.query(Material).filter(Material.id == config.skin_material_id).first()
        skin_area = area * 2 * waste_factor
        materials.append({
            "material_name": skin_mat.name if skin_mat else "Piel OSB",
            "category": "sheathing",
            "quantity": round(skin_area, 3),
            "unit": "m2",
            "cost_per_unit": skin_mat.cost_per_unit if skin_mat and skin_mat.cost_per_unit else 0,
            "cost": round(skin_area * (skin_mat.cost_per_unit if skin_mat and skin_mat.cost_per_unit else 0), 2),
        })

        # Core material
        core_mat = self.db.query(Material).filter(Material.id == config.core_material_id).first()
        core_volume = area * (config.core_thickness_mm / 1000) * waste_factor
        materials.append({
            "material_name": core_mat.name if core_mat else "Núcleo aislante",
            "category": "insulation",
            "quantity": round(core_volume, 4),
            "unit": "m3",
            "cost_per_unit": core_mat.cost_per_unit if core_mat and core_mat.cost_per_unit else 0,
            "cost": round(core_volume * (core_mat.cost_per_unit if core_mat and core_mat.cost_per_unit else 0), 2),
        })

        # Fasteners
        perimeter_mm = 2 * (instance.width_mm + instance.height_mm)
        screw_count = math.ceil(perimeter_mm * SIP_FASTENER_RULES["screws_per_mm_perimeter"])
        screw_count += math.ceil(area * 1e6 * SIP_FASTENER_RULES["screws_per_mm_attachment"] / 150)
        screw_count = math.ceil(screw_count * waste_factor)
        materials.append({
            "material_name": "Tornillos SIP",
            "category": "hardware",
            "quantity": screw_count,
            "unit": "unidades",
            "cost_per_unit": 0.15,
            "cost": round(screw_count * 0.15, 2),
        })

        # Adhesive
        adhesive_kg = area * SIP_FASTENER_RULES["adhesive_kg_per_m2"] * waste_factor
        materials.append({
            "material_name": "Adhesivo SIP",
            "category": "adhesive",
            "quantity": round(adhesive_kg, 3),
            "unit": "kg",
            "cost_per_unit": 8.0,
            "cost": round(adhesive_kg * 8.0, 2),
        })

        return materials

    def _wood_frame_panel_materials(self, instance: PanelInstance, waste_factor: float) -> list[dict]:
        """Calculate materials for a wood frame panel from its framing members."""
        materials_agg: dict[str, dict] = {}

        members = (
            self.db.query(FramingMember)
            .filter(FramingMember.panel_instance_id == instance.id)
            .all()
        )

        for member in members:
            key = f"Madera {member.lumber_size} ({member.member_type})"
            length_m = (member.length_mm / 1000) * member.quantity * waste_factor

            if key in materials_agg:
                materials_agg[key]["quantity"] += length_m
            else:
                materials_agg[key] = {
                    "material_name": f"Madera {member.lumber_size}",
                    "category": "structural",
                    "quantity": length_m,
                    "unit": "ml",
                    "cost_per_unit": 3.0,  # default cost per linear meter
                    "cost": 0,
                }

        # Calculate costs and round
        for mat in materials_agg.values():
            mat["quantity"] = round(mat["quantity"], 2)
            mat["cost"] = round(mat["quantity"] * mat["cost_per_unit"], 2)

        # Nails
        nail_count = 0
        for member in members:
            if member.member_type in ("stud", "king_stud", "jack_stud", "cripple_stud"):
                nail_count += 3 * member.quantity  # 3 nails per connection
        nail_count = math.ceil(nail_count * waste_factor)

        materials_list = list(materials_agg.values())
        if nail_count > 0:
            materials_list.append({
                "material_name": "Clavos entramado",
                "category": "hardware",
                "quantity": nail_count,
                "unit": "unidades",
                "cost_per_unit": 0.05,
                "cost": round(nail_count * 0.05, 2),
            })

        return materials_list
